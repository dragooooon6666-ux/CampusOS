"""
文件内容提取模块

针对大学生常用文件类型提取文本内容，供 AI 命名使用。
支持：docx / pdf / xlsx / txt / md / 纯文本
"""

import logging
from pathlib import Path

# ── 提取结果 ──────────────────────────────────────────────

MAX_PREVIEW_CHARS = 800   # 发给 AI 的最大字数
TITLE_LINES = 6           # 取前 N 行作为标题候选区
SIGNATURE_LINES = 4       # 取最后 N 行作为落款候选区


class ContentPreview:
    """提取后的内容预览，包含标题区和落款区"""

    def __init__(self, title_section: str = "", signature_section: str = ""):
        self.title_section = title_section
        self.signature_section = signature_section
        self.full_text = ""

    @property
    def ai_context(self) -> str:
        """拼接成适合发给 AI 的上下文文本"""
        parts = []
        if self.title_section:
            parts.append(f"【文件开头内容】\n{self.title_section}")
        if self.signature_section:
            parts.append(f"【文件末尾内容（落款/署名）】\n{self.signature_section}")
        return "\n\n".join(parts) if parts else ""

    def __bool__(self):
        return bool(self.ai_context)


# ── 各类型提取器 ───────────────────────────────────────────

def _extract_docx(file_path: Path) -> ContentPreview:
    """从 Word 文档提取文本段落"""
    import docx

    doc = docx.Document(str(file_path))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

    if not paragraphs:
        return ContentPreview()

    return _build_preview(paragraphs)


def _extract_pdf(file_path: Path) -> ContentPreview:
    """从 PDF 提取文本"""
    from PyPDF2 import PdfReader

    reader = PdfReader(str(file_path))
    all_lines: list[str] = []

    for page in reader.pages[:3]:  # 最多读 3 页
        text = page.extract_text()
        if text:
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            all_lines.extend(lines)
        if sum(len(l) for l in all_lines) > MAX_PREVIEW_CHARS * 2:
            break

    if not all_lines:
        return ContentPreview()

    return _build_preview(all_lines)


def _extract_xlsx(file_path: Path) -> ContentPreview:
    """从 Excel 提取单元格文本（前 20 行）"""
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    lines: list[str] = []

    for sheet_name in wb.sheetnames[:2]:  # 最多读 2 个工作表
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > 20:
                break
            row_text = " ".join(str(c) for c in row if c is not None).strip()
            if row_text:
                lines.append(row_text)

    wb.close()
    return _build_preview(lines)


def _extract_plain(file_path: Path) -> ContentPreview:
    """从纯文本/txt/md 读取内容"""
    try:
        text = file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        try:
            text = file_path.read_text(encoding="gbk")
        except UnicodeDecodeError:
            return ContentPreview()

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return _build_preview(lines)


def _extract_nothing(file_path: Path) -> ContentPreview:
    """无法提取内容的文件类型"""
    return ContentPreview()


# ── 通用构建 ───────────────────────────────────────────────

def _build_preview(lines: list[str]) -> ContentPreview:
    """从行列表中分离标题区和落款区"""
    if not lines:
        return ContentPreview()

    preview = ContentPreview()

    # 标题区：前 N 行（截断到合理长度）
    title_chars = 0
    title_lines = []
    for line in lines[:TITLE_LINES]:
        title_lines.append(line)
        title_chars += len(line)
        if title_chars >= MAX_PREVIEW_CHARS // 2:
            break
    preview.title_section = "\n".join(title_lines)[:MAX_PREVIEW_CHARS]

    # 落款区：最后 N 行
    if len(lines) > TITLE_LINES + 2:
        sig = lines[-SIGNATURE_LINES:]
        preview.signature_section = "\n".join(sig)[:MAX_PREVIEW_CHARS // 2]

    preview.full_text = "\n".join(lines)
    return preview


# ── 类型 → 提取器映射 ──────────────────────────────────────

EXTRACTOR_MAP = {
    ".docx": _extract_docx,
    ".doc": _extract_docx,
    ".pdf": _extract_pdf,
    ".xlsx": _extract_xlsx,
    ".xls": _extract_xlsx,
    ".csv": _extract_plain,
    ".txt": _extract_plain,
    ".md": _extract_plain,
    ".markdown": _extract_plain,
    ".log": _extract_plain,
}

FALLBACK_EXTRACTOR = _extract_nothing


# ── 公开 API ───────────────────────────────────────────────

def extract_content_preview(file_path: Path) -> ContentPreview:
    """
    根据文件扩展名提取内容预览。
    返回 ContentPreview，包含标题区和落款区文本。
    """
    extension = file_path.suffix.lower()
    extractor = EXTRACTOR_MAP.get(extension, FALLBACK_EXTRACTOR)

    try:
        preview = extractor(file_path)
    except Exception as exc:
        logging.warning("内容提取失败 [%s]: %s", file_path.name, exc)
        preview = ContentPreview()

    # 如果标题区为空但文件名有意义，用文件名作为 fallback 上下文
    if not preview.ai_context:
        stem = file_path.stem.strip()
        if stem and len(stem) >= 2:
            preview.title_section = f"（文件名为：{stem}）"

    return preview
